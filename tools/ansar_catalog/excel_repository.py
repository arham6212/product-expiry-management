from __future__ import annotations

from copy import copy
from datetime import date, datetime, timezone
import os
from pathlib import Path
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .config import PRODUCT_COLUMNS, RUN_LOG_COLUMNS
from .normalizer import identity_text


TEXT_FIELDS = {
    "id", "run_id", "global_barcode", "ansar_sku", "first_seen_at",
    "last_seen_at", "started_at", "completed_at",
}
DATE_FIELDS = {"first_seen_at", "last_seen_at", "started_at", "completed_at"}
FORMULA_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")


def _cell_value(header: str, value):
    if value is None:
        return None
    if header in DATE_FIELDS and isinstance(value, (date, datetime)):
        if isinstance(value, date) and not isinstance(value, datetime):
            value = datetime.combine(value, datetime.min.time())
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    if header in TEXT_FIELDS:
        return str(value).lstrip("'")
    return value


def _rows(sheet, headers: list[str]) -> list[dict]:
    actual = [str(cell.value or "") for cell in sheet[1][:len(headers)]]
    if actual != headers:
        raise RuntimeError(f"Unexpected workbook headers in {sheet.title}: {actual}")
    rows: list[dict] = []
    for values in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        if not any(value is not None and value != "" for value in values):
            continue
        row = {header: _cell_value(header, values[index]) for index, header in enumerate(headers)}
        barcode = row.get("global_barcode")
        if isinstance(barcode, (int, float)):
            length = {"GTIN-8": 8, "UPC-A": 12, "EAN-13": 13, "GTIN-14": 14}.get(row.get("barcode_type"))
            row["global_barcode"] = str(int(barcode)).zfill(length or 1)
        sku = row.get("ansar_sku")
        if isinstance(sku, (int, float)):
            row["ansar_sku"] = str(int(sku))
        rows.append(row)
    return rows


def _style_header(sheet, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="0F766E")
    font = Font(bold=True, color="FFFFFF")
    for index, header in enumerate(headers, 1):
        cell = sheet.cell(1, index, header)
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def _new_workbook():
    workbook = Workbook()
    products = workbook.active
    products.title = "Products"
    logs = workbook.create_sheet("Run_Log")
    _style_header(products, PRODUCT_COLUMNS)
    _style_header(logs, RUN_LOG_COLUMNS)
    return workbook


def _copy_previous_style(sheet, row_number: int, column_count: int) -> None:
    if row_number <= 2:
        return
    for column in range(1, column_count + 1):
        source = sheet.cell(row_number - 1, column)
        target = sheet.cell(row_number, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format


def _write_row(sheet, row_number: int, headers: list[str], values: dict) -> None:
    _copy_previous_style(sheet, row_number, len(headers))
    for column, header in enumerate(headers, 1):
        value = values.get(header)
        cell = sheet.cell(row_number, column, value)
        if header in TEXT_FIELDS:
            cell.number_format = "@"
            if value is not None:
                cell.value = str(value)


def _update_table(sheet, table_name: str, column_count: int) -> None:
    ref = f"A1:{get_column_letter(column_count)}{max(2, sheet.max_row)}"
    tables = list(sheet.tables.values())
    if tables:
        tables[0].ref = ref
    else:
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)
    sheet.auto_filter.ref = ref


class ExcelRepository:
    def __init__(self, workbook_path: Path) -> None:
        self.path = workbook_path

    def _load(self):
        return load_workbook(self.path) if self.path.exists() else _new_workbook()

    def _save_atomic(self, workbook) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        fd, temporary = tempfile.mkstemp(prefix=f".{self.path.name}.", suffix=".xlsx", dir=self.path.parent)
        os.close(fd)
        try:
            workbook.save(temporary)
            os.replace(temporary, self.path)
        finally:
            Path(temporary).unlink(missing_ok=True)

    def snapshot(self) -> dict:
        if not self.path.exists():
            return {"products": [], "run_log": []}
        workbook = load_workbook(self.path, data_only=False)
        return {
            "products": _rows(workbook["Products"], PRODUCT_COLUMNS),
            "run_log": _rows(workbook["Run_Log"], RUN_LOG_COLUMNS),
        }

    @staticmethod
    def _keys(row: dict) -> tuple[str | None, str]:
        barcode = str(row.get("global_barcode") or "").strip() or None
        fallback = "|".join([
            identity_text(str(row.get("canonical_name") or row.get("source_display_name") or "")),
            str(row.get("ansar_sku") or "").strip(),
            str(row.get("source_product_url") or "").strip(),
        ])
        return barcode, fallback

    def append(self, products: list[dict], run_log: dict) -> tuple[int, int]:
        existing = self.snapshot()
        barcodes: set[str] = set()
        fallback_keys: set[str] = set()
        for row in existing["products"]:
            barcode, fallback = self._keys(row)
            if barcode:
                barcodes.add(barcode)
            fallback_keys.add(fallback)

        accepted: list[dict] = []
        dropped = 0
        for row in products:
            barcode, fallback = self._keys(row)
            if (barcode and barcode in barcodes) or fallback in fallback_keys:
                dropped += 1
                continue
            accepted.append({column: row.get(column) for column in PRODUCT_COLUMNS})
            if barcode:
                barcodes.add(barcode)
            fallback_keys.add(fallback)

        log = {column: run_log.get(column) for column in RUN_LOG_COLUMNS}
        log["new_products_added"] = len(accepted)
        log["duplicates_skipped"] = int(log.get("duplicates_skipped") or 0) + dropped

        workbook = self._load()
        product_sheet = workbook["Products"]
        log_sheet = workbook["Run_Log"]
        for row in accepted:
            _write_row(product_sheet, product_sheet.max_row + 1, PRODUCT_COLUMNS, row)
        _write_row(log_sheet, log_sheet.max_row + 1, RUN_LOG_COLUMNS, log)
        _update_table(product_sheet, "ProductsTable", len(PRODUCT_COLUMNS))
        _update_table(log_sheet, "RunLogTable", len(RUN_LOG_COLUMNS))
        self._save_atomic(workbook)
        return len(accepted), dropped

    def verify(self, preview_dir: Path | None = None) -> dict:
        del preview_dir
        workbook = load_workbook(self.path, data_only=False)
        products = _rows(workbook["Products"], PRODUCT_COLUMNS)
        logs = _rows(workbook["Run_Log"], RUN_LOG_COLUMNS)
        barcodes = [str(row.get("global_barcode") or "") for row in products if row.get("global_barcode")]
        urls = [str(row.get("source_product_url") or "") for row in products if row.get("source_product_url")]
        duplicate_barcodes = sorted({value for value in barcodes if barcodes.count(value) > 1})
        duplicate_urls = sorted({value for value in urls if urls.count(value) > 1})
        formula_errors: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and any(error in value for error in FORMULA_ERRORS):
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}:{value}")
        return {
            "products": len(products),
            "run_log_rows": len(logs),
            "valid_global_barcodes": sum(bool(row.get("barcode_valid")) for row in products),
            "invalid_or_internal_identifiers": sum(not bool(row.get("barcode_valid")) for row in products),
            "duplicate_global_barcodes": duplicate_barcodes,
            "duplicate_source_urls": duplicate_urls,
            "headers_valid": True,
            "formula_error_scan": formula_errors,
        }

    def patch_products(self, patches: list[dict]) -> int:
        if not patches:
            return 0
        workbook = self._load()
        sheet = workbook["Products"]
        header_index = {cell.value: cell.column for cell in sheet[1]}
        rows_by_id = {str(sheet.cell(row, header_index["id"]).value): row for row in range(2, sheet.max_row + 1)}
        for patch in patches:
            row_number = rows_by_id.get(str(patch["id"]))
            if not row_number:
                raise RuntimeError(f"Product row not found for correction: {patch['id']}")
            for header, value in patch["values"].items():
                sheet.cell(row_number, header_index[header], value)
        self._save_atomic(workbook)
        return len(patches)
